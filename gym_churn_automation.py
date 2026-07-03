import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 1. 파일 이름 정의
INPUT_FILE = "members.xlsx"
OUTPUT_FILE = "gym_churn_report.xlsx"
EMAIL_FILE = "retention_emails.txt"

def create_mock_data():
    """데이터가 없을 경우 자동으로 생성하는 초기 회원 엑셀 파일 생성기"""
    print(f"[{INPUT_FILE}] 파일이 존재하지 않아 가상 회원 데이터를 생성합니다...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "회원명단"

    # 헤더 정의
    headers = ["회원번호", "이름", "이메일", "회원권 종류", "가입일", "만료일", "마지막 방문일", "누적 출석수"]
    ws.append(headers)

    # 샘플 데이터 (이름, 이메일, 회원권, 가입일, 만료일, 마지막방문일, 누적출석)
    # 오늘 날짜를 기준으로 작성하여 위험군 계산이 명확하게 작동하도록 함
    today = datetime.date.today()
    
    mock_members = [
        (1, "김지성", "jisung@example.com", "12개월 회원권", today - datetime.timedelta(days=180), today + datetime.timedelta(days=185), today - datetime.timedelta(days=1), 120),
        (2, "박민아", "mina@example.com", "3개월 회원권", today - datetime.timedelta(days=75), today + datetime.timedelta(days=15), today - datetime.timedelta(days=5), 22),
        (3, "이현우", "hyunwoo@example.com", "1개월 회원권", today - datetime.timedelta(days=25), today + datetime.timedelta(days=5), today - datetime.timedelta(days=18), 4),
        (4, "정소민", "somin@example.com", "6개월 회원권", today - datetime.timedelta(days=150), today + datetime.timedelta(days=30), today - datetime.timedelta(days=21), 34),
        (5, "최주환", "juhwan@example.com", "12개월 회원권", today - datetime.timedelta(days=330), today - datetime.timedelta(days=2), today - datetime.timedelta(days=14), 95),
        (6, "윤서준", "seojun@example.com", "3개월 회원권", today - datetime.timedelta(days=10), today + datetime.timedelta(days=80), today - datetime.timedelta(days=2), 5),
    ]

    for member in mock_members:
        ws.append(member)

    # 날짜 서식 설정
    for row in range(2, len(mock_members) + 2):
        for col in [5, 6, 7]:
            ws.cell(row=row, column=col).number_format = 'YYYY-MM-DD'

    wb.save(INPUT_FILE)
    print(f"[{INPUT_FILE}] 생성 완료!")


def analyze_churn_risk():
    """회원 데이터를 읽어 이탈 위험도를 분석하고 스타일링된 엑셀 보고서 및 개인화 이메일을 생성하는 핵심 함수"""
    if not os.path.exists(INPUT_FILE):
        create_mock_data()

    print("\n--- 헬스장 고객 이탈 방지 프로그램(파이썬 일잘러 버전) 구동 ---")
    
    # 엑셀 파일 로드
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # 분석결과를 저장할 새 시트 생성
    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    report_ws.title = "이탈위험 분석결과"

    # 새 헤더 추가 (기존 헤더 + 분석 필드)
    headers = ["회원번호", "이름", "이메일", "회원권 종류", "만료일", "마지막 방문일", "미출석 일수", "남은 만료일", "이탈 위험도", "이탈방지 조치"]
    report_ws.append(headers)

    # 디자인용 스타일 정의
    font_title = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Malgun Gothic", size=10)
    font_bold = Font(name="Malgun Gothic", size=10, bold=True)
    
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # 다크 그레이
    
    # 위험도별 색상 정의 (소프트한 파스텔톤 적용)
    fill_high = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid") # 연빨강
    fill_medium = PatternFill(start_color="FDE047", end_color="FDE047", fill_type="solid") # 연노랑
    fill_low = PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid") # 연초록

    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    today = datetime.date.today()
    email_drafts = []

    # 데이터 분석 및 행 작성
    row_idx = 2
    for r in range(2, ws.max_row + 1):
        # 기존 데이터 파싱
        m_id = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        email = ws.cell(row=r, column=3).value
        m_type = ws.cell(row=r, column=4).value
        
        # 문자열로 들어오는 경우 대비하여 날짜 객체로 파싱
        end_date = ws.cell(row=r, column=6).value
        last_visit = ws.cell(row=r, column=7).value

        if isinstance(end_date, str):
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        elif isinstance(end_date, datetime.datetime):
            end_date = end_date.date()

        if isinstance(last_visit, str):
            last_visit = datetime.datetime.strptime(last_visit, "%Y-%m-%d").date()
        elif isinstance(last_visit, datetime.datetime):
            last_visit = last_visit.date()

        # 미출석 일수 및 만료일 계산
        days_since_visit = (today - last_visit).days
        days_until_expire = (end_date - today).days

        # 위험 등급 판정 알고리즘
        risk = "LOW"
        action = "정상 출석 중"
        
        if days_since_visit >= 14 or days_until_expire <= 7:
            risk = "HIGH"
            action = "1:1 PT 쿠폰 및 15% 할인팩 발송 대상"
        elif days_since_visit >= 7 or days_until_expire <= 30:
            risk = "MEDIUM"
            action = "안내 문자 및 재등록 안내 대상"

        # 결과 데이터 쓰기
        row_data = [m_id, name, email, m_type, end_date, last_visit, days_since_visit, days_until_expire, risk, action]
        report_ws.append(row_data)

        # 서식 지정
        for c in range(1, len(row_data) + 1):
            cell = report_ws.cell(row=row_idx, column=c)
            cell.font = font_data
            cell.border = border_thin
            if c in [1, 5, 6, 7, 8, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

        # 위험 등급 열 색상 강조
        risk_cell = report_ws.cell(row=row_idx, column=9)
        risk_cell.font = font_bold
        if risk == "HIGH":
            risk_cell.fill = fill_high
        elif risk == "MEDIUM":
            risk_cell.fill = fill_medium
        else:
            risk_cell.fill = fill_low

        # 개인화 이탈 방지 이메일 본문 작성 (HIGH / MEDIUM 위험군 대상)
        if risk in ["HIGH", "MEDIUM"]:
            email_content = f"""==================================================
수신자: {name} 회원님 ({email})
발신자: FIT-SHIELD 센터 관리자
제목: [FIT-SHIELD] {name} 회원님, 다시 달릴 준비가 되셨나요? 특별 혜택을 드립니다!

안녕하세요, {name} 회원님! FIT-SHIELD 헬스장입니다.

최근 회원님의 바쁜 일상으로 인해 운동 방문이 뜸해지신 것을 확인했습니다.
회원님께서 처음에 세우셨던 건강 목표를 포기하지 않으시도록 저희가 특별한 복귀 선물을 준비했습니다.

🎁 회원님만을 위한 특별 복귀 쿠폰팩:
1. 1:1 개인 PT 1회 무료 체험권 (바코드 확인 즉시 예약 가능)
2. 회원권 재등록 및 연장 시 적용되는 15% 특별 추가 할인 혜택

* 마지막 방문일: {last_visit} (미방문 일수: {days_since_visit}일)
* 이용권 만료일: {end_date} (남은 기간: {days_until_expire}일)

인포데스크에 본 메시지를 보여주시거나 아래의 임시 쿠폰 코드를 제시해 주세요:
[ 쿠폰 코드: RECOVERY-{name.upper()}-{m_id} ]

회원님의 건강한 라이프스타일을 FIT-SHIELD가 항상 응원합니다. 센터에서 기다리겠습니다!
==================================================\n"""
            email_drafts.append(email_content)

        row_idx += 1

    # 헤더 행 디자인 적용
    for col in range(1, len(headers) + 1):
        cell = report_ws.cell(row=1, column=col)
        cell.font = font_title
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # 열 너비 자동 조절
    for col in report_ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # 한글 인코딩 길이 보정을 위해 대략적인 문자열 길이 측정
            val_str = str(cell.value or '')
            val_len = len(val_str.encode('utf-8'))
            if val_len > max_len:
                max_len = val_len
        # 조금 넉넉하게 너비 설정
        report_ws.column_dimensions[col_letter].width = max(max_len // 2 + 3, 12)

    # 보고서 파일 저장
    report_wb.save(OUTPUT_FILE)
    print(f"[OK] 분석 결과 엑셀 파일 저장 완료: [{OUTPUT_FILE}] (위험도별 색상 구분 완료)")

    # 이메일 텍스트 저장
    with open(EMAIL_FILE, "w", encoding="utf-8") as f:
        f.writelines(email_drafts)
    print(f"[OK] 이탈 방지 개인화 이메일 초안 작성 완료: [{EMAIL_FILE}] (총 {len(email_drafts)}건)")

    # 요약 출력
    total_analyzed = len(report_ws['I']) - 1
    high_count = sum(1 for cell in report_ws['I'] if cell.value == 'HIGH')
    medium_count = sum(1 for cell in report_ws['I'] if cell.value == 'MEDIUM')
    
    print("\n--- 분석 결과 요약 ---")
    print(f"총 분석 회원 수: {total_analyzed}명")
    print(f"[HIGH] 위험군: {high_count}명 (즉시 쿠폰팩 제공 대상)")
    print(f"[MEDIUM] 경고군: {medium_count}명 (만료 전 연장 권장 대상)")
    print(f"안내메일/문자 초안이 '{EMAIL_FILE}'에 저장되었습니다.")
    print("---------------------------------------------")

if __name__ == "__main__":
    analyze_churn_risk()
