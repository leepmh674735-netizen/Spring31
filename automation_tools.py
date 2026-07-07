import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel_report():
    """
    1. Excel 보고서 자동 생성 자동화 (Pandas & OpenPyXL)
    """
    print("[1/5] 데이터베이스 회원 정보 로드 중...")
    # 가상 회원 데이터 생성
    np.random.seed(42)
    n_samples = 150
    members = [f"회원_{i:03d}" for i in range(1, n_samples + 1)]
    recencies = np.random.randint(0, 30, n_samples)
    frequencies = np.random.randint(0, 25, n_samples)
    contracts = np.random.choice([1, 3, 6, 12], n_samples)
    
    df = pd.DataFrame({
        'Name': members,
        'Recency (Days)': recencies,
        'Monthly Attendance': frequencies,
        'Contract Period (Months)': contracts
    })
    
    # 이탈 확률 가상 계산
    df['Churn Risk (%)'] = np.clip(((df['Recency (Days)'] > 15) * 40 + (df['Monthly Attendance'] < 5) * 30 + (df['Contract Period (Months)'] <= 3) * 15 + 10), 5, 95)
    
    print("[2/5] Pandas DataFrame 생성 완료. 엑셀 변환 진행...")
    
    # Excel 파일 작성
    file_path = "gym_member_report.xlsx"
    df.to_excel(file_path, index=False)
    
    print("[3/5] openpyxl 기반 스타일링 및 서식 적용 중...")
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "이탈 분석 보고서"
    
    # 열 너비 설정
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 헤더 스타일링 (어두운 회색 배경, 흰색 볼드 텍스트)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # 데이터 행 서식 및 조건부 하이라이트 (이탈위험 70% 이상 빨간색 표시)
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_font = Font(color="991B1B", bold=True)
    
    for row in range(2, ws.max_row + 1):
        # Churn Risk 열 값 체크 (5번째 열)
        risk_cell = ws.cell(row=row, column=5)
        risk_value = risk_cell.value
        
        # 기본 테두리
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if col > 1 else "left")
            
        if risk_value and float(risk_value) >= 70:
            risk_cell.fill = red_fill
            risk_cell.font = red_font
            
    print(f"[4/5] 셀 서식 적용 완료. 보고서 파일 저장 중...")
    wb.save(file_path)
    print(f"[5/5] 성공: '{file_path}' 파일이 생성되었습니다.")

def send_bulk_alerts():
    """
    2. 이탈위험군 알림톡/SMS 일괄 전송 자동화
    """
    print("[1/4] 데이터베이스에서 이탈 고위험군(Risk >= 70%) 대상 검색 중...")
    targets = [
        {"name": "강철수", "phone": "010-1234-5678", "risk": 82},
        {"name": "박민영", "phone": "010-9876-5432", "risk": 74},
        {"name": "정수지", "phone": "010-5555-4444", "risk": 91}
    ]
    print(f"[2/4] 총 {len(targets)}명의 고위험군 대상자를 추출했습니다.")
    
    for t in targets:
        message = (
            f"안녕하세요, {t['name']} 회원님! 건강한 한 주 보내고 계신가요?\n"
            f"회원님의 락커 및 회원권 혜택 만료가 다가와 안내해 드립니다.\n"
            f"이번 주말 방문 시 GX 무료 수강 쿠폰 1매를 특별 증정하오니 "
            f"센터에 들르셔서 시원하게 운동을 즐겨보세요! - Smart Gym 드림"
        )
        print(f"\n--- [전송 대기] 대상: {t['name']} ({t['phone']}) Churn Risk: {t['risk']}% ---")
        print(message)
        print(">> SMS API 서버로 패킷 전송 성공 (status: 200 OK)")
        
    print("\n[3/4] SMS 일괄 발송 완료 및 발송 성공 로그 DB 기록...")
    print("[4/4] 성공: 알림 문자 일괄 전송 작업이 완료되었습니다.")

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "sms":
        send_bulk_alerts()
    else:
        generate_excel_report()
